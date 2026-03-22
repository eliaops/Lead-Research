"""Agent sync API — endpoints for local agents to communicate with the cloud.

Security:
  - Authenticated via X-Agent-Key header (AGENT_API_KEY env var)
  - Separate from the SCRAPER_API_KEY used by the web app
  - No secrets exposed in responses

Endpoints:
  GET  /api/agent/jobs               — fetch pending jobs for local_connector sources
  POST /api/agent/jobs/create        — create pending jobs
  POST /api/agent/jobs/{id}/status   — update job status
  POST /api/agent/opportunities      — upload batch of normalized opportunities
  POST /api/agent/documents          — upload document metadata for an opportunity
  GET  /api/agent/pending-documents  — list high-relevance opps needing document download
  POST /api/agent/upload-documents   — upload actual document files + trigger analysis
"""

from __future__ import annotations

import io
import json
from datetime import date, datetime, timezone
from decimal import Decimal

from fastapi import APIRouter, Depends, File, Form, Header, HTTPException, UploadFile, status
from pydantic import BaseModel, Field
from sqlalchemy import text
from typing import Any

from src.core.config import settings
from src.core.database import get_db_session
from src.core.logging import get_logger
from src.models.opportunity import (
    AgentDocumentUpload,
    AgentJobResponse,
    AgentOpportunityUpload,
    AgentStatusUpdate,
    RunStatus,
    TriggerType,
)
from src.utils.dedup import check_duplicate, check_source_duplicate, generate_fingerprint
from src.utils.scorer import score_opportunity

logger = get_logger(__name__)
router = APIRouter(prefix="/api/agent", tags=["agent"])


class _SafeEncoder(json.JSONEncoder):
    def default(self, o: object) -> object:
        if isinstance(o, (datetime, date)):
            return o.isoformat()
        if isinstance(o, Decimal):
            return float(o)
        return super().default(o)


def _json(obj: object) -> str | None:
    if obj is None:
        return None
    return json.dumps(obj, cls=_SafeEncoder)


def verify_agent_key(x_agent_key: str = Header(...)) -> str:
    key = settings.AGENT_API_KEY or settings.SCRAPER_API_KEY
    if not key:
        raise HTTPException(status.HTTP_500_INTERNAL_SERVER_ERROR, "AGENT_API_KEY not configured")
    if x_agent_key != key:
        raise HTTPException(status.HTTP_401_UNAUTHORIZED, "Invalid agent key")
    return x_agent_key


# ─── GET /api/agent/jobs ────────────────────────────────────


@router.get("/jobs", dependencies=[Depends(verify_agent_key)])
async def get_pending_jobs() -> list[AgentJobResponse]:
    """Return pending crawl runs for local_connector sources."""
    session = get_db_session()
    try:
        rows = session.execute(
            text("""
                SELECT sr.id AS run_id, s.id AS source_id, s.name AS source_name,
                       s.base_url, s.crawl_config, s.access_mode
                FROM source_runs sr
                JOIN sources s ON sr.source_id = s.id
                WHERE sr.status = 'pending'
                  AND s.access_mode = 'local_connector'
                  AND s.is_active = true
                ORDER BY sr.created_at ASC
                LIMIT 10
            """)
        ).fetchall()

        jobs = []
        for r in rows:
            jobs.append(AgentJobResponse(
                run_id=str(r.run_id),
                source_id=str(r.source_id),
                source_name=r.source_name,
                base_url=r.base_url,
                crawl_config=r.crawl_config if r.crawl_config else {},
                access_mode=r.access_mode or "local_connector",
            ))
        return jobs
    finally:
        session.close()


# ─── POST /api/agent/jobs/create ────────────────────────────


class CreateJobRequest(BaseModel):
    source_name: str | None = None


@router.post("/jobs/create", dependencies=[Depends(verify_agent_key)])
async def create_agent_job(req: CreateJobRequest | None = None) -> list[AgentJobResponse]:
    """Create pending crawl runs for all local_connector sources.

    The local agent calls this to request work, or it can be called by the
    cloud scheduler to pre-create jobs.
    """
    session = get_db_session()
    try:
        where = "s.access_mode = 'local_connector' AND s.is_active = true"
        params: dict[str, Any] = {}
        if req and req.source_name:
            where += " AND s.name = :name"
            params["name"] = req.source_name

        sources = session.execute(
            text(f"SELECT id, name, base_url, crawl_config, access_mode FROM sources s WHERE {where}"),
            params,
        ).fetchall()

        jobs = []
        for s in sources:
            # Don't create duplicate pending jobs
            existing = session.execute(
                text("""
                    SELECT id FROM source_runs
                    WHERE source_id = :sid AND status IN ('pending', 'running')
                    LIMIT 1
                """),
                {"sid": str(s.id)},
            ).fetchone()
            if existing:
                jobs.append(AgentJobResponse(
                    run_id=str(existing.id),
                    source_id=str(s.id),
                    source_name=s.name,
                    base_url=s.base_url,
                    crawl_config=s.crawl_config if s.crawl_config else {},
                    access_mode=s.access_mode or "local_connector",
                ))
                continue

            row = session.execute(
                text("""
                    INSERT INTO source_runs (source_id, status, triggered_by)
                    VALUES (:sid, 'pending', 'local_agent')
                    RETURNING id
                """),
                {"sid": str(s.id)},
            ).fetchone()
            session.commit()

            jobs.append(AgentJobResponse(
                run_id=str(row.id),  # type: ignore
                source_id=str(s.id),
                source_name=s.name,
                base_url=s.base_url,
                crawl_config=s.crawl_config if s.crawl_config else {},
                access_mode=s.access_mode or "local_connector",
            ))

        logger.info("Created %d agent jobs", len(jobs))
        return jobs
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()


# ─── POST /api/agent/jobs/{run_id}/status ───────────────────


@router.post("/jobs/{run_id}/status", dependencies=[Depends(verify_agent_key)])
async def update_job_status(run_id: str, update: AgentStatusUpdate) -> dict:
    """Update a crawl run's status and stats from the local agent."""
    session = get_db_session()
    try:
        now = datetime.now(timezone.utc)

        started_clause = ""
        if update.status == RunStatus.RUNNING:
            started_clause = ", started_at = :now"

        completed_clause = ""
        if update.status in (RunStatus.COMPLETED, RunStatus.FAILED, RunStatus.CANCELLED):
            completed_clause = ", completed_at = :now, duration_ms = EXTRACT(EPOCH FROM (:now - COALESCE(started_at, :now)))::int * 1000"

        session.execute(
            text(f"""
                UPDATE source_runs SET
                    status = :status,
                    pages_crawled = :pages,
                    opportunities_found = :found,
                    opportunities_created = :created,
                    opportunities_updated = :updated,
                    opportunities_skipped = :skipped,
                    error_message = :error_msg,
                    error_details = :error_details
                    {started_clause}
                    {completed_clause}
                WHERE id = :run_id
            """),
            {
                "run_id": run_id,
                "status": update.status.value,
                "pages": update.pages_crawled,
                "found": update.opportunities_found,
                "created": update.opportunities_created,
                "updated": update.opportunities_updated,
                "skipped": update.opportunities_skipped,
                "error_msg": update.error_message,
                "error_details": json.dumps(update.error_details) if update.error_details else None,
                "now": now,
            },
        )

        if update.status in (RunStatus.COMPLETED, RunStatus.FAILED):
            session.execute(
                text("""
                    UPDATE sources SET
                        last_crawled_at = :now,
                        last_run_status = :status
                    WHERE id = (SELECT source_id FROM source_runs WHERE id = :run_id)
                """),
                {"now": now, "status": update.status.value, "run_id": run_id},
            )

        session.commit()
        logger.info("Agent updated run %s → %s", run_id, update.status.value)
        return {"status": "ok", "run_id": run_id}
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()


# ─── POST /api/agent/opportunities ──────────────────────────


@router.post("/opportunities", dependencies=[Depends(verify_agent_key)])
async def upload_opportunities(batch: AgentOpportunityUpload) -> dict:
    """Receive a batch of opportunities from the local agent.

    Each opportunity is normalized, scored, deduped, and inserted — exactly
    the same pipeline as cloud-executed crawls.
    """
    session = get_db_session()
    created = 0
    updated = 0
    skipped = 0
    errors: list[str] = []

    try:
        for opp in batch.opportunities:
            try:
                opp.source_id = batch.source_id
                opp.source_run_id = batch.run_id

                # Score
                desc = opp.description_full or opp.description_summary or ""
                score, breakdown = score_opportunity(
                    title=opp.title, description=desc,
                    org_type=None, project_type=opp.project_type,
                    category=opp.category, source_fit_score=85,
                )
                opp.relevance_score = score
                opp.relevance_breakdown = breakdown
                opp.relevance_bucket = breakdown.get("relevance_bucket", "irrelevant")
                opp.keywords_matched = (
                    breakdown.get("primary_matches", [])
                    + breakdown.get("secondary_matches", [])
                    + breakdown.get("contextual_matches", [])
                )
                opp.negative_keywords = breakdown.get("negative_matches", [])
                opp.industry_tags = breakdown.get("industry_tags", [])

                # Fingerprint
                closing_str = str(opp.closing_date) if opp.closing_date else ""
                opp.fingerprint = generate_fingerprint(
                    opp.title, opp.organization_name or "", closing_str, opp.source_url,
                )

                # Dedup
                if opp.external_id:
                    existing_id = check_source_duplicate(session, opp.source_id, opp.external_id)
                    if existing_id:
                        _update_opp(session, existing_id, opp)
                        updated += 1
                        continue

                existing_id = check_duplicate(session, opp.fingerprint)
                if existing_id:
                    skipped += 1
                    continue

                _insert_opp(session, opp)
                created += 1

                if (opp.relevance_score or 0) >= 80 and opp.external_id:
                    try:
                        row = session.execute(
                            text("SELECT id FROM opportunities WHERE external_id = :eid AND source_id = :sid LIMIT 1"),
                            {"eid": opp.external_id, "sid": opp.source_id},
                        ).fetchone()
                        if row:
                            from src.tasks.auto_analyze import auto_analyze_opportunity
                            auto_analyze_opportunity.apply_async(args=[str(row.id)], countdown=60)
                            logger.info("Dispatched auto-analysis for agent-uploaded opp: %s", opp.title[:60])
                    except Exception as aa_exc:
                        logger.warning("Failed to dispatch auto-analysis for agent opp: %s", aa_exc)

            except Exception as exc:
                errors.append(f"{opp.title[:60]}: {exc}")
                logger.exception("Agent upload: failed to process %s", opp.title[:60])

        session.commit()
        logger.info(
            "Agent upload for run %s: created=%d updated=%d skipped=%d errors=%d",
            batch.run_id, created, updated, skipped, len(errors),
        )
        return {
            "status": "ok",
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": len(errors),
        }
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()


# ─── POST /api/agent/documents ──────────────────────────────


@router.post("/documents", dependencies=[Depends(verify_agent_key)])
async def upload_documents(batch: AgentDocumentUpload) -> dict:
    """Receive document metadata from the local agent."""
    session = get_db_session()
    try:
        opp_row = session.execute(
            text("""
                SELECT id FROM opportunities
                WHERE source_id = :sid AND external_id = :eid
                LIMIT 1
            """),
            {"sid": batch.source_id, "eid": batch.opportunity_external_id},
        ).fetchone()

        if not opp_row:
            return {"status": "skipped", "reason": "opportunity not found"}

        opp_id = str(opp_row.id)
        inserted = 0

        for doc in batch.documents:
            existing = session.execute(
                text("SELECT id FROM opportunity_documents WHERE opportunity_id = :oid AND url = :url LIMIT 1"),
                {"oid": opp_id, "url": doc.get("url", "")},
            ).fetchone()
            if existing:
                continue

            session.execute(
                text("""
                    INSERT INTO opportunity_documents (
                        opportunity_id, title, url, file_type, file_size_bytes,
                        page_count, doc_category
                    ) VALUES (
                        :oid, :title, :url, :ft, :size, :pages, :cat
                    )
                """),
                {
                    "oid": opp_id,
                    "title": doc.get("name", "Document"),
                    "url": doc.get("url", ""),
                    "ft": doc.get("file_type"),
                    "size": doc.get("file_size_bytes"),
                    "pages": doc.get("page_count"),
                    "cat": doc.get("doc_category"),
                },
            )
            inserted += 1

        if inserted > 0:
            session.execute(
                text("UPDATE opportunities SET has_documents = true WHERE id = :oid"),
                {"oid": opp_id},
            )

        session.commit()
        return {"status": "ok", "documents_inserted": inserted}
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()


# ─── GET /api/agent/pending-documents ────────────────────────


@router.get("/pending-documents", dependencies=[Depends(verify_agent_key)])
async def get_pending_documents(
    source_name: str | None = None,
    min_score: int = 80,
    limit: int = 20,
) -> list[dict]:
    """Return high-relevance opportunities that need document download.

    Used by local agents (e.g. BT Agent) to know which opportunities
    to fetch bid documents for.
    """
    session = get_db_session()
    try:
        params: dict[str, Any] = {"min_score": min_score, "limit": limit}
        source_filter = ""
        if source_name:
            source_filter = "AND s.name = :source_name"
            params["source_name"] = source_name

        rows = session.execute(
            text(f"""
                SELECT o.id, o.external_id, o.title, o.title_zh,
                       o.relevance_score, o.source_url, o.closing_date,
                       o.has_documents,
                       org.name as organization_name,
                       s.name as source_name
                FROM opportunities o
                JOIN sources s ON o.source_id = s.id
                LEFT JOIN organizations org ON o.organization_id = org.id
                WHERE o.relevance_score >= :min_score
                  AND o.has_documents = false
                  AND o.status = 'open'
                  {source_filter}
                ORDER BY o.relevance_score DESC, o.created_at DESC
                LIMIT :limit
            """),
            params,
        ).fetchall()

        results = []
        for r in rows:
            results.append({
                "opportunity_id": str(r.id),
                "external_id": r.external_id,
                "title": r.title_zh or r.title,
                "relevance_score": r.relevance_score,
                "source_url": r.source_url,
                "closing_date": r.closing_date.isoformat() if r.closing_date else None,
                "organization_name": r.organization_name,
                "source_name": r.source_name,
            })

        logger.info("Pending documents query: %d results (source=%s, min_score=%d)",
                     len(results), source_name, min_score)
        return results
    finally:
        session.close()


# ─── POST /api/agent/upload-documents ────────────────────────


_ALLOWED_EXTENSIONS = {"pdf", "docx", "doc", "txt", "xlsx", "xls", "csv"}
_MAX_FILE_SIZE = 25 * 1024 * 1024  # 25MB


@router.post("/upload-documents", dependencies=[Depends(verify_agent_key)])
async def upload_document_files(
    opportunity_id: str = Form(...),
    files: list[UploadFile] = File(...),
    trigger_analysis: bool = Form(True),
) -> dict:
    """Upload actual document files for an opportunity and optionally trigger AI analysis.

    Accepts multipart/form-data with 1-10 files. Extracts text from each,
    stores in opportunity_documents, and dispatches auto_analyze_opportunity.
    """
    if not files or len(files) > 10:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Must upload 1-10 files")

    session = get_db_session()
    try:
        opp = session.execute(
            text("SELECT id, title FROM opportunities WHERE id = :id"),
            {"id": opportunity_id},
        ).fetchone()
        if not opp:
            raise HTTPException(status.HTTP_404_NOT_FOUND, f"Opportunity {opportunity_id} not found")

        inserted = 0
        for upload in files:
            filename = upload.filename or "document"
            ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
            if ext not in _ALLOWED_EXTENSIONS:
                logger.warning("Skipping unsupported file type: %s", filename)
                continue

            content = await upload.read()
            if len(content) > _MAX_FILE_SIZE:
                logger.warning("File too large (%d bytes): %s", len(content), filename)
                continue

            extracted_text = _extract_text_from_bytes(content, ext, filename)

            session.execute(
                text("""
                    INSERT INTO opportunity_documents (
                        opportunity_id, title, url, file_type,
                        file_size_bytes, doc_category,
                        extracted_text, text_extracted
                    ) VALUES (
                        :opp_id, :title, :url, :ft,
                        :size, 'agent_upload',
                        :text, :extracted
                    )
                """),
                {
                    "opp_id": opportunity_id,
                    "title": filename,
                    "url": f"agent-upload://{filename}",
                    "ft": ext,
                    "size": len(content),
                    "text": extracted_text,
                    "extracted": bool(extracted_text),
                },
            )
            inserted += 1
            logger.info("Stored uploaded document: %s (%d bytes, %d chars text)",
                        filename, len(content), len(extracted_text or ""))

        if inserted > 0:
            session.execute(
                text("UPDATE opportunities SET has_documents = true WHERE id = :id"),
                {"id": opportunity_id},
            )
            session.commit()

            if trigger_analysis:
                try:
                    from src.tasks.auto_analyze import auto_analyze_opportunity
                    auto_analyze_opportunity.apply_async(args=[opportunity_id], countdown=10)
                    logger.info("Dispatched auto-analysis after document upload for %s", opportunity_id)
                except Exception as exc:
                    logger.warning("Failed to dispatch analysis: %s", exc)
        else:
            session.commit()

        return {
            "status": "ok",
            "documents_stored": inserted,
            "analysis_triggered": trigger_analysis and inserted > 0,
        }
    except HTTPException:
        raise
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()


def _extract_text_from_bytes(content: bytes, ext: str, filename: str) -> str | None:
    """Extract text from file bytes. Returns None on failure."""
    try:
        if ext == "pdf":
            from pypdf import PdfReader
            reader = PdfReader(io.BytesIO(content))
            pages = []
            for page in reader.pages:
                t = page.extract_text()
                if t:
                    pages.append(t)
            return "\n\n".join(pages)[:200_000] if pages else None

        if ext in ("docx", "doc"):
            from docx import Document
            doc = Document(io.BytesIO(content))
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            return "\n".join(paragraphs)[:200_000] if paragraphs else None

        if ext == "txt":
            for enc in ("utf-8", "latin-1", "cp1252"):
                try:
                    return content.decode(enc)[:200_000]
                except UnicodeDecodeError:
                    continue
            return None

        if ext in ("xlsx", "xls"):
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            rows_text = []
            for ws in wb.worksheets[:5]:
                for row in ws.iter_rows(max_row=500, values_only=True):
                    cells = [str(c) for c in row if c is not None]
                    if cells:
                        rows_text.append("\t".join(cells))
            return "\n".join(rows_text)[:200_000] if rows_text else None

        if ext == "csv":
            text_content = content.decode("utf-8", errors="replace")
            return text_content[:200_000]

        return None
    except Exception as exc:
        logger.warning("Text extraction failed for %s: %s", filename, exc)
        return None


# ─── Internal helpers ───────────────────────────────────────


def _insert_opp(session: Any, opp: Any) -> None:
    session.execute(
        text("""
            INSERT INTO opportunities (
                source_id, source_run_id, external_id,
                title, description_summary, description_full,
                status, country, region, city, location_raw,
                posted_date, closing_date, project_type, category,
                solicitation_number, estimated_value, currency,
                contact_name, contact_email, contact_phone,
                source_url, has_documents,
                keywords_matched, negative_keywords, relevance_score,
                relevance_bucket, relevance_breakdown, industry_tags,
                ingestion_mode, raw_data, fingerprint, updated_at
            ) VALUES (
                :source_id, :source_run_id, :external_id,
                :title, :description_summary, :description_full,
                :status, :country, :region, :city, :location_raw,
                :posted_date, :closing_date, :project_type, :category,
                :solicitation_number, :estimated_value, :currency,
                :contact_name, :contact_email, :contact_phone,
                :source_url, :has_documents,
                :keywords_matched, :negative_keywords, :relevance_score,
                :relevance_bucket, :relevance_breakdown, :industry_tags,
                'live', :raw_data, :fingerprint, NOW()
            )
        """),
        {
            "source_id": opp.source_id,
            "source_run_id": opp.source_run_id,
            "external_id": opp.external_id,
            "title": opp.title,
            "description_summary": opp.description_summary,
            "description_full": opp.description_full,
            "status": opp.status.value if opp.status else "unknown",
            "country": opp.country,
            "region": opp.region,
            "city": opp.city,
            "location_raw": opp.location_raw,
            "posted_date": opp.posted_date,
            "closing_date": opp.closing_date,
            "project_type": opp.project_type,
            "category": opp.category,
            "solicitation_number": opp.solicitation_number,
            "estimated_value": float(opp.estimated_value) if opp.estimated_value else None,
            "currency": opp.currency,
            "contact_name": opp.contact_name,
            "contact_email": opp.contact_email,
            "contact_phone": opp.contact_phone,
            "source_url": opp.source_url,
            "has_documents": opp.has_documents,
            "keywords_matched": opp.keywords_matched,
            "negative_keywords": opp.negative_keywords,
            "relevance_score": opp.relevance_score,
            "relevance_bucket": opp.relevance_bucket,
            "relevance_breakdown": _json(opp.relevance_breakdown),
            "industry_tags": opp.industry_tags,
            "raw_data": _json(opp.raw_data),
            "fingerprint": opp.fingerprint,
        },
    )


def _update_opp(session: Any, opp_id: str, opp: Any) -> None:
    session.execute(
        text("""
            UPDATE opportunities SET
                source_run_id = :source_run_id,
                title = :title,
                description_summary = COALESCE(:description_summary, description_summary),
                description_full = COALESCE(:description_full, description_full),
                status = :status,
                closing_date = COALESCE(:closing_date, closing_date),
                keywords_matched = :keywords_matched,
                negative_keywords = :negative_keywords,
                relevance_score = :relevance_score,
                relevance_bucket = :relevance_bucket,
                relevance_breakdown = :relevance_breakdown,
                industry_tags = :industry_tags,
                updated_at = NOW()
            WHERE id = :id
        """),
        {
            "id": opp_id,
            "source_run_id": opp.source_run_id,
            "title": opp.title,
            "description_summary": opp.description_summary,
            "description_full": opp.description_full,
            "status": opp.status.value if opp.status else "unknown",
            "closing_date": opp.closing_date,
            "keywords_matched": opp.keywords_matched,
            "negative_keywords": opp.negative_keywords,
            "relevance_score": opp.relevance_score,
            "relevance_bucket": opp.relevance_bucket,
            "relevance_breakdown": _json(opp.relevance_breakdown),
            "industry_tags": opp.industry_tags,
        },
    )
