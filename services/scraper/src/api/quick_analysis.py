"""Analysis API — on-demand AI analysis for opportunities.

Two endpoints:
  POST /api/analysis/upload-and-analyze — Upload documents, get full Markdown report
  POST /api/analysis/mini-summary      — Lightweight 2-3 sentence AI assessment
"""

from __future__ import annotations

import io
import json
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

import requests as http_requests
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile, status
from sqlalchemy import text

from src.api.auth import verify_api_key
from src.core.config import settings
from src.core.database import get_db_session
from src.core.logging import get_logger
from src.intelligence.analyzer import TenderAnalyzer

logger = get_logger(__name__)
router = APIRouter(prefix="/api/analysis", tags=["analysis"])

def _estimate_cost(model: str, prompt_tokens: int, completion_tokens: int) -> float:
    rates = {"gpt-4o": (2.50, 10.00), "gpt-4o-mini": (0.15, 0.60)}
    inp, out = rates.get(model, (2.50, 10.00))
    return round((prompt_tokens / 1_000_000) * inp + (completion_tokens / 1_000_000) * out, 6)


def _check_budget(session: Any) -> tuple[bool, str]:
    try:
        today = session.execute(
            text("SELECT COALESCE(SUM(estimated_cost_usd), 0) as t FROM ai_usage_log WHERE created_at >= CURRENT_DATE"),
        ).fetchone()
        daily = float(today.t) if today else 0.0
        if daily >= settings.AI_DAILY_BUDGET_USD:
            return False, f"日预算已用完 (${daily:.2f}/${settings.AI_DAILY_BUDGET_USD:.2f})"

        month = session.execute(
            text("SELECT COALESCE(SUM(estimated_cost_usd), 0) as t FROM ai_usage_log WHERE created_at >= date_trunc('month', CURRENT_DATE)"),
        ).fetchone()
        monthly = float(month.t) if month else 0.0
        if monthly >= settings.AI_MONTHLY_BUDGET_USD:
            return False, f"月预算已用完 (${monthly:.2f}/${settings.AI_MONTHLY_BUDGET_USD:.2f})"
        return True, ""
    except Exception as exc:
        logger.warning("Budget check failed (allowing): %s", exc)
        return True, ""


def _record_usage(session: Any, opp_id: str, model: str, mode: str,
                  prompt_tok: int, completion_tok: int, cost: float) -> None:
    try:
        session.execute(
            text("""INSERT INTO ai_usage_log (
                opportunity_id, model, analysis_mode, prompt_tokens, completion_tokens,
                total_tokens, estimated_cost_usd, created_at
            ) VALUES (:oid, :model, :mode, :pt, :ct, :tt, :cost, :ts)"""),
            {"oid": opp_id, "model": model, "mode": mode, "pt": prompt_tok,
             "ct": completion_tok, "tt": prompt_tok + completion_tok,
             "cost": cost, "ts": datetime.now(timezone.utc)},
        )
    except Exception as exc:
        logger.warning("Failed to record AI usage: %s", exc)


# ─── Text Extraction ─────────────────────────────────────────


def _extract_text_from_bytes(content: bytes, file_type: str) -> str:
    ft = file_type.lower().strip(".")
    if ft == "pdf":
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(content))
        parts = [p.extract_text() or "" for p in reader.pages]
        return "\n\n".join(p for p in parts if p)
    if ft in ("docx", "doc"):
        from docx import Document
        doc = Document(io.BytesIO(content))
        return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())
    if ft in ("xlsx", "xls"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        parts: list[str] = []
        for sheet in wb.worksheets[:10]:
            for row in sheet.iter_rows(max_row=1000, values_only=True):
                cells = [str(c) for c in row if c is not None]
                if cells:
                    parts.append("\t".join(cells))
        wb.close()
        return "\n".join(parts)
    if ft == "csv":
        return content.decode("utf-8", errors="replace")
    if ft == "txt":
        return content.decode("utf-8", errors="replace")
    return ""


# ─── Upload & Analyze ────────────────────────────────────────


_UPLOAD_MAX_FILES = 10
_UPLOAD_MAX_SIZE = 25 * 1024 * 1024
_SUPPORTED_EXTS = {"pdf", "docx", "doc", "txt", "xlsx", "xls", "csv"}


@router.post("/upload-and-analyze", dependencies=[Depends(verify_api_key)])
async def upload_and_analyze(
    files: list[UploadFile] = File(...),
    opportunity_id: str | None = Form(None),
) -> dict:
    """Upload up to 10 documents and run deep AI analysis (gpt-4o, 16K tokens, ~$0.20-$0.50).

    Returns Markdown report with cost info. Budget cap: $5/analysis.
    """
    if len(files) > _UPLOAD_MAX_FILES:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, f"最多上传 {_UPLOAD_MAX_FILES} 个文件")

    document_texts: dict[str, str] = {}
    stored_files: list[str] = []

    for f in files:
        if not f.filename:
            continue
        content = await f.read()
        if len(content) > _UPLOAD_MAX_SIZE:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, f"文件 {f.filename} 超过 25MB 限制")
        ext = (f.filename.rsplit(".", 1)[-1] if "." in f.filename else "").lower()
        if ext not in _SUPPORTED_EXTS:
            raise HTTPException(status.HTTP_400_BAD_REQUEST,
                                f"不支持的文件类型: {f.filename}（支持: {', '.join(_SUPPORTED_EXTS)}）")
        try:
            extracted = _extract_text_from_bytes(content, ext)
            if extracted and len(extracted.strip()) > 10:
                document_texts[f.filename] = extracted
                stored_files.append(f.filename)
        except Exception as exc:
            logger.warning("Failed to extract text from %s: %s", f.filename, exc)

    if not document_texts:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "未能从上传的文件中提取到有效文本")

    session = get_db_session()
    try:
        budget_ok, budget_msg = _check_budget(session)
        if not budget_ok:
            return {"status": "budget_exceeded", "message": budget_msg}

        opp_title = "上传文档分析"
        organization = None
        location = None
        closing_date = None
        source_name = "手动上传"
        description = ""
        country = None
        solicitation_number = None

        if opportunity_id:
            opp = session.execute(
                text("""
                    SELECT o.id, o.title, o.description_summary, o.description_full,
                           o.country, o.region, o.city, o.closing_date,
                           o.solicitation_number,
                           s.name as source_name, org.name as organization_name
                    FROM opportunities o
                    LEFT JOIN sources s ON o.source_id = s.id
                    LEFT JOIN organizations org ON o.organization_id = org.id
                    WHERE o.id = :id
                """),
                {"id": opportunity_id},
            ).fetchone()

            if opp:
                opp_title = opp.title
                organization = opp.organization_name
                lp = [p for p in [opp.city, opp.region, opp.country] if p]
                location = ", ".join(lp) if lp else None
                closing_date = str(opp.closing_date) if opp.closing_date else None
                source_name = opp.source_name or "Unknown"
                description = opp.description_full or opp.description_summary or ""
                country = opp.country
                solicitation_number = opp.solicitation_number

            # Also include any previously uploaded/extracted docs
            existing_docs = session.execute(
                text("""
                    SELECT title, extracted_text FROM opportunity_documents
                    WHERE opportunity_id = :id AND text_extracted = true
                      AND extracted_text IS NOT NULL AND LENGTH(extracted_text) > 10
                    LIMIT 10
                """),
                {"id": opportunity_id},
            ).fetchall()
            for i, doc in enumerate(existing_docs):
                key = doc.title or f"existing_doc_{i}"
                if key not in document_texts:
                    document_texts[key] = doc.extracted_text

            # Store uploaded files as opportunity_documents
            for f_name in stored_files:
                session.execute(
                    text("""
                        INSERT INTO opportunity_documents (
                            opportunity_id, title, url, file_type,
                            doc_category, extracted_text, text_extracted
                        ) VALUES (:oid, :title, :url, :ft, 'tender_document', :text, true)
                    """),
                    {
                        "oid": opportunity_id,
                        "title": f_name,
                        "url": f"upload://{f_name}",
                        "ft": f_name.rsplit(".", 1)[-1].lower() if "." in f_name else "unknown",
                        "text": document_texts.get(f_name, "")[:200000],
                    },
                )
            if stored_files:
                session.execute(
                    text("UPDATE opportunities SET has_documents = true WHERE id = :id"),
                    {"id": opportunity_id},
                )

        analyzer = TenderAnalyzer(model="gpt-4o", max_tokens=16000)
        result = analyzer.analyze(
            title=opp_title,
            organization=organization,
            location=location,
            closing_date=closing_date,
            source=source_name,
            description=description,
            document_texts=document_texts,
            solicitation_number=solicitation_number,
            country=country,
        )

        report_md = result.get("report_markdown", "")
        model = result.get("model", "gpt-4o")
        prompt_tok = result.get("prompt_tokens", 0)
        completion_tok = result.get("completion_tokens", 0)
        actual_cost = result.get("estimated_cost_usd", 0.0)
        is_fallback = result.get("fallback", False)
        now = datetime.now(timezone.utc)

        if not is_fallback:
            cost = actual_cost or _estimate_cost(model, prompt_tok, completion_tok)
            _record_usage(session, opportunity_id or "upload", model, "upload_deep", prompt_tok, completion_tok, cost)

        if opportunity_id:
            summary_json = json.dumps({
                "report_markdown": report_md,
                "report_version": "5.0",
                "documents_analyzed": stored_files,
                "analysis_cost_usd": actual_cost,
                "model": model,
            })

            existing = session.execute(
                text("SELECT id FROM tender_intelligence WHERE opportunity_id = :id"),
                {"id": opportunity_id},
            ).fetchone()

            params = {
                "opp_id": opportunity_id,
                "overview": report_md[:500],
                "summary": summary_json,
                "model": model,
                "mode": "upload_deep",
                "status": "completed" if not is_fallback else "fallback_only",
                "now": now,
            }

            if existing:
                session.execute(text("""
                    UPDATE tender_intelligence SET
                        project_overview = :overview,
                        intelligence_summary = :summary,
                        analysis_model = :model,
                        analysis_mode = :mode,
                        analysis_status = :status,
                        analyzed_at = :now, updated_at = :now
                    WHERE opportunity_id = :opp_id
                """), params)
            else:
                session.execute(text("""
                    INSERT INTO tender_intelligence (
                        opportunity_id, project_overview,
                        intelligence_summary, analysis_model, analysis_mode,
                        analysis_status, analyzed_at, updated_at
                    ) VALUES (
                        :opp_id, :overview,
                        :summary, :model, :mode,
                        :status, :now, :now
                    )
                """), params)

            session.commit()

        return {
            "status": "completed",
            "opportunity_id": opportunity_id,
            "report_markdown": report_md,
            "documents_analyzed": stored_files,
            "model": model,
            "cost_usd": round(actual_cost, 4),
            "tokens": {"prompt": prompt_tok, "completion": completion_tok},
        }

    except HTTPException:
        raise
    except Exception as exc:
        session.rollback()
        logger.exception("Upload analysis failed")
        return {"status": "error", "message": str(exc)}
    finally:
        session.close()


# ─── Mini Summary ─────────────────────────────────────────────


from pydantic import BaseModel


class MiniSummaryRequest(BaseModel):
    opportunity_id: str


@router.post("/mini-summary", dependencies=[Depends(verify_api_key)])
async def generate_mini_summary(req: MiniSummaryRequest) -> dict:
    """Generate a lightweight AI summary (2-3 sentences) for an opportunity."""
    session = get_db_session()
    try:
        opp = session.execute(
            text("""
                SELECT o.id, o.title, o.description_summary, o.description_full,
                       o.closing_date, o.business_fit_explanation,
                       org.name as organization_name,
                       o.region, o.city, o.country
                FROM opportunities o
                LEFT JOIN organizations org ON o.organization_id = org.id
                WHERE o.id = :id
            """),
            {"id": req.opportunity_id},
        ).fetchone()

        if not opp:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Opportunity not found")

        if opp.business_fit_explanation and len(opp.business_fit_explanation) > 30:
            return {"status": "exists", "summary": opp.business_fit_explanation}

        description = opp.description_full or opp.description_summary or ""
        lp = [p for p in [opp.city, opp.region, opp.country] if p]

        summary = TenderAnalyzer.generate_mini_summary(
            title=opp.title,
            description=description,
            organization=opp.organization_name,
            location=", ".join(lp) if lp else None,
            closing_date=str(opp.closing_date) if opp.closing_date else None,
        )

        if summary:
            session.execute(
                text("UPDATE opportunities SET business_fit_explanation = :s, updated_at = NOW() WHERE id = :id"),
                {"s": summary, "id": req.opportunity_id},
            )
            session.commit()

            cost = _estimate_cost("gpt-4o-mini", 500, 150)
            _record_usage(session, req.opportunity_id, "gpt-4o-mini", "mini_summary", 500, 150, cost)

            return {"status": "generated", "summary": summary}

        return {"status": "failed", "summary": None}

    except HTTPException:
        raise
    except Exception as exc:
        session.rollback()
        logger.exception("Mini summary failed")
        return {"status": "error", "message": str(exc)}
    finally:
        session.close()


# ─── Analysis Status ──────────────────────────────────────────


@router.get("/status/{opportunity_id}", dependencies=[Depends(verify_api_key)])
async def analysis_status(opportunity_id: str) -> dict:
    session = get_db_session()
    try:
        row = session.execute(
            text("""SELECT id, analysis_model, analyzed_at, intelligence_summary
                    FROM tender_intelligence WHERE opportunity_id = :id"""),
            {"id": opportunity_id},
        ).fetchone()
        if row:
            summary = row.intelligence_summary
            if isinstance(summary, str):
                try:
                    summary = json.loads(summary)
                except Exception:
                    summary = {}
            has_markdown = bool(summary.get("report_markdown")) if isinstance(summary, dict) else False
            return {
                "exists": True,
                "intel_id": str(row.id),
                "model": row.analysis_model,
                "analyzed_at": row.analyzed_at.isoformat() if row.analyzed_at else None,
                "has_markdown_report": has_markdown,
            }
        return {"exists": False}
    finally:
        session.close()


@router.post("/cleanup-old-data", dependencies=[Depends(verify_api_key)])
async def cleanup_old_analysis_data() -> dict:
    """Remove all old v3 JSON analysis data. One-time migration endpoint."""
    session = get_db_session()
    try:
        intel_count = session.execute(text("SELECT COUNT(*) FROM tender_intelligence")).scalar() or 0
        if intel_count > 0:
            session.execute(text("DELETE FROM tender_intelligence"))

        biz_count = session.execute(
            text("UPDATE opportunities SET business_fit_explanation = NULL WHERE business_fit_explanation IS NOT NULL")
        ).rowcount

        agent_docs = session.execute(
            text("DELETE FROM opportunity_documents WHERE url LIKE 'agent-upload://%' OR url LIKE 'upload://%'")
        ).rowcount

        session.commit()

        return {
            "status": "ok",
            "deleted_analyses": intel_count,
            "cleared_biz_fit": biz_count,
            "deleted_agent_docs": agent_docs,
        }
    except Exception as exc:
        session.rollback()
        logger.exception("Cleanup failed")
        return {"status": "error", "message": str(exc)}
    finally:
        session.close()
